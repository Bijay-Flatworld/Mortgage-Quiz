from flask import Flask, render_template, request, redirect, url_for, session
import random, os
from openpyxl import load_workbook, Workbook

app = Flask(__name__)
app.secret_key = "quiz_secret"

EXCEL_PATH = os.path.join(os.getcwd(), "mortgage_quiz.xlsx")

questions = [
    {"q": "What is a mortgage?", "options": ["Loan", "Savings", "Tax", "Insurance"]},
    {"q": "What is EMI?", "options": ["Monthly Payment", "Credit Score", "Tax", "Penalty"]},
    {"q": "What does LTV stand for?", "options": ["Loan to Value", "Lease to Value", "Long Term Value", "Legal Tax Value"]},
    {"q": "Which factor affects interest rate?", "options": ["Credit Score", "Loan Amount", "Tenure", "All of these"]},
    {"q": "What is foreclosure?", "options": ["Loan default", "Extra EMI", "Interest waiver", "Tax rebate"]},
    {"q": "What is the typical tenure of a mortgage loan?", "options": ["1 year", "5-30 years", "50 years", "Unlimited"]},
    {"q": "What is the principal amount?", "options": ["Original loan amount", "Interest only", "Tax amount", "Monthly EMI"]},
    {"q": "What does amortization mean?", "options": ["Gradual repayment", "Penalty", "Default", "Tax credit"]},
    {"q": "What is refinancing?", "options": ["Replacing loan with new one", "Skipping EMI", "Government subsidy", "Tax filing"]},
    {"q": "Which is a secured loan?", "options": ["Mortgage", "Credit Card", "Personal Loan", "Overdraft"]},
    {"q": "Who is the borrower?", "options": ["Person taking loan", "Bank", "Government", "Employer"]},
    {"q": "Who is the lender?", "options": ["Bank/Financial Institution", "Borrower", "Government", "Employer"]},
    {"q": "What is the purpose of collateral?", "options": ["Secure the loan", "Pay interest", "Reduce EMI", "Increase tenure"]},
    {"q": "What is prepayment?", "options": ["Pay loan early", "Miss EMI", "Increase loan", "Take tax rebate"]},
    {"q": "What is the penalty for default?", "options": ["Late fee/foreclosure", "Interest reduction", "No penalty", "Tenure increase"]},
    {"q": "What is fixed interest rate?", "options": ["Same rate entire loan", "Rate changes", "Rate waived", "No interest"]},
    {"q": "What is floating interest rate?", "options": ["Rate varies with market", "Fixed forever", "No interest", "Government set"]},
    {"q": "What is down payment?", "options": ["Initial upfront payment", "Monthly EMI", "Tax", "Interest"]},
    {"q": "What is equity in home loan?", "options": ["Ownership value", "Loan balance", "Interest", "Penalty"]},
    {"q": "What is credit score used for?", "options": ["Loan eligibility", "Movie rating", "Tax filing", "Electricity bill"]},
    {"q": "What is a co-borrower?", "options": ["Joint loan applicant", "Only guarantor", "Bank staff", "Tax officer"]},
    {"q": "What is loan default?", "options": ["Failure to repay", "Extra EMI", "Tenure extension", "Tax rebate"]},
    {"q": "What is an escrow account?", "options": ["Account for property tax & insurance", "Salary account", "Savings", "Trading account"]},
    {"q": "What is pre-approval?", "options": ["Conditional loan approval", "Final approval", "Tax rebate", "Penalty waiver"]},
    {"q": "What is a mortgage broker?", "options": ["Middleman between borrower & lender", "Bank manager", "Auditor", "Lawyer"]},
    {"q": "What is balloon payment?", "options": ["Large final payment", "Missed EMI", "Tax credit", "Foreclosure"]},
    {"q": "What is interest-only loan?", "options": ["Pay interest first", "Pay principal only", "Pay nothing", "Pay all upfront"]},
    {"q": "What is the annual percentage rate (APR)?", "options": ["True cost of loan incl. fees", "Only principal", "Tax amount", "Discount rate"]},
    {"q": "What is the grace period?", "options": ["Extra time before EMI", "Loan default", "Tax filing time", "No EMI forever"]},
    {"q": "What is a loan modification?", "options": ["Change terms of existing loan", "New loan", "Cancel loan", "Government subsidy"]},
    {"q": "What is negative amortization?", "options": ["Loan balance increases", "Loan balance decreases", "Tax rebate", "Penalty waiver"]},
    {"q": "What is a mortgage-backed security?", "options": ["Investment backed by mortgages", "Insurance policy", "Tax bond", "Savings account"]},
    {"q": "What is PMI (Private Mortgage Insurance)?", "options": ["Insurance for lender if borrower defaults", "Life insurance", "Tax scheme", "Interest waiver"]},
    {"q": "What is DTI ratio?", "options": ["Debt to Income ratio", "Deposit to Investment", "Dividend to Interest", "None"]},
    {"q": "What is a guarantor?", "options": ["Person guaranteeing repayment", "Borrower", "Bank", "Tax officer"]},
    {"q": "What is repossession?", "options": ["Lender takes back property", "Extra EMI", "Tax rebate", "Loan top-up"]},
    {"q": "What is a reverse mortgage?", "options": ["Loan for seniors against home equity", "Loan for students", "Short-term loan", "No loan"]},
    {"q": "What is tenure?", "options": ["Duration of loan", "Monthly EMI", "Interest type", "Tax amount"]},
]

def get_excel():
    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        wb.save(EXCEL_PATH)
    return load_workbook(EXCEL_PATH)

@app.route("/", methods=["GET", "POST"])
def entry():
    if request.method == "POST":
        session["name"] = request.form["username"].strip()
        session["role"] = request.form["role"]
        session["qset"] = random.sample(questions, 20)
        session["qindex"] = 0
        return redirect(url_for("quiz"))
    return render_template("entry.html")

@app.route("/quiz", methods=["GET", "POST"])
def quiz():
    if "qindex" not in session:
        return redirect(url_for("entry"))

    if request.method == "POST":
        answer = request.form.get("answer", "No Answer")
        q = session["qset"][session["qindex"]]["q"]

        wb = get_excel()
        if session["name"] not in wb.sheetnames:
            ws = wb.create_sheet(session["name"])
            ws.append(["Question", "Answer"])
        ws = wb[session["name"]]
        ws.append([q, answer])
        wb.save(EXCEL_PATH)

        session["qindex"] += 1
        if session["qindex"] >= 20:
            return redirect(url_for("thankyou"))

    qnum = session["qindex"] + 1
    qdata = session["qset"][session["qindex"]]
    return render_template("quiz.html", qnum=qnum, total=20, question=qdata)

@app.route("/thankyou")
def thankyou():
    return render_template("thankyou.html")

if __name__ == "__main__":
    app.run(debug=True)
